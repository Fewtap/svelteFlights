import json
import os
from time import sleep
import openpyxl
from openpyxl import Workbook
import datetime
from pocketbase import PocketBase
from pocketbase.client import FileUpload
import requests
from pocketbase.models import Record


# make a global variable for the old rooms


def updatefiles(date):
    client = PocketBase("http://176.58.101.163:8080")
    fileexists: bool = False

    # get tomorrows date

    # convert the date to a iso string

    querystring = date.isoformat()

    result: list[Record] = client.collection("sheet").get_full_list(
        query_params={"filter": f'date ~ "{querystring}"'}
    )

    if len(result) > 0:
        fileexists = True
        print("File already exists")

    rutecolumn = 1
    buscolumn = 2
    roomcolumn = 3
    amountcolumn = 5

    # get absolute path to the file
    path = os.path.abspath("transfer.xlsx")
    print(path)

    wb = openpyxl.open(path, read_only=False)

    def getflights(collection: str):
        # convert the date to a YYYY-MM-DD string
        datestring = date.strftime("%Y-%m-%d")
        url = f'http://176.58.101.163:8080/api/collections/{collection}/records?filter=(planned~"{datestring}")'
        response = requests.get(url)
        data = response.json()
        return data

    def getrooms(hash: str, datestring: str = None):
        if datestring is not None:
            response = client.collection("rooms").get_full_list(
                query_params={"filter": f'flighthash="{hash}" && date ~ "{datestring}"'}
            )
        else:
            response = client.collection("rooms").get_full_list(
                query_params={"filter": f'flighthash="{hash}"'}
            )
            print("Response without datestring:")
            print(response)

        return response

    departureswithrooms = []
    for flight in getflights("departures")["items"]:
        hasrooms = flight["hasrooms"]

        if hasrooms == True:
            print("Adding flight: " + flight["rute"] + " to list")
            departureswithrooms.append(flight)

    arrivalswithrooms = []
    for flight in getflights("arrivals")["items"]:
        hasrooms = flight["hasrooms"]

        if hasrooms == True:
            print("Adding flight: " + flight["rute"] + " to list")
            arrivalswithrooms.append(flight)

    # sort the lists by the planned key
    departureswithrooms = sorted(departureswithrooms, key=lambda x: x["planned"])
    arrivalswithrooms = sorted(arrivalswithrooms, key=lambda x: x["planned"])

    # create the excel file
    sheet = wb.active

    # add the headers

    for i in range(5, 17):
        sheet.cell(row=i, column=amountcolumn).value = ""
        sheet.cell(row=i, column=roomcolumn).value = ""
        sheet.cell(row=i, column=buscolumn).value = ""
        sheet.cell(row=i, column=rutecolumn).value = ""

    for i in range(21, 32):
        sheet.cell(row=i, column=amountcolumn).value = ""
        sheet.cell(row=i, column=roomcolumn).value = ""
        sheet.cell(row=i, column=buscolumn).value = ""
        sheet.cell(row=i, column=rutecolumn).value = ""

    row = 5

    for i in range(len(departureswithrooms)):
        totalpeople = 0
        rooms = getrooms(departureswithrooms[i]["flighthash"])
        for room in rooms:
            print(room.roomnumber + " " + str(room.amount))
            totalpeople += room.amount
        departuretimestring = departureswithrooms[i]["planned"]
        departuretime = datetime.datetime.fromisoformat(
            departuretimestring[:-1] + "+00:00"
        )
        departuretime = departuretime - datetime.timedelta(hours=3)
        busdeparturetime = departuretime - datetime.timedelta(minutes=90)
        busdeparturetimestring = busdeparturetime.strftime("%H:%M")
        flight = departureswithrooms[i]
        sheet.cell(row=row, column=rutecolumn).value = flight["rute"]
        sheet.cell(row=row, column=buscolumn).value = busdeparturetimestring
        roomstring = ""
        for room in rooms:
            roomstring += room.roomnumber + ";"

        print("Roomstring Arrivals: " + roomstring)
        sheet.cell(row=row, column=roomcolumn).value = roomstring
        sheet.cell(row=row, column=amountcolumn).value = totalpeople

        row += 1

    row = 21

    for i in range(len(arrivalswithrooms)):
        totalpeople = 0
        rooms = getrooms(arrivalswithrooms[i]["flighthash"])
        for room in rooms:
            totalpeople += room.amount
        departuretimestring = arrivalswithrooms[i]["planned"]
        departuretime = datetime.datetime.fromisoformat(
            departuretimestring[:-1] + "+00:00"
        )
        departuretime = departuretime - datetime.timedelta(hours=3)
        # convert the time to local time

        timestring = departuretime.strftime("%H:%M")
        flight = arrivalswithrooms[i]
        sheet.cell(row=row, column=rutecolumn).value = flight["rute"]
        sheet.cell(row=row, column=buscolumn).value = timestring
        roomstring = ""
        for room in rooms:
            roomstring += room.roomnumber + ";"
        print("Roomstring Arrivals: " + roomstring)
        sheet.cell(row=row, column=roomcolumn).value = roomstring
        sheet.cell(row=row, column=amountcolumn).value = totalpeople

        row += 1

    datestring = date.strftime("%Y-%m-%d")
    sheet.cell(row=1, column=1).value = datestring

    roomswithouttransfers = getrooms("", datestring)

    # filter out the rooms where the key departure is false and put them in a list called arrivalswithouttransfers
    arrivalswithouttransfers = []
    departureswithouttransfers = []
    for room in roomswithouttransfers:
        if room.departure == False:
            arrivalswithouttransfers.append(room)
        else:
            departureswithouttransfers.append(room)

    departurestring = ""
    departureamount = 0
    for room in departureswithouttransfers:
        departurestring += room.roomnumber + ";"
        departureamount += room.amount

    arrivalstring = ""
    arrivalamount = 0
    for room in arrivalswithouttransfers:
        arrivalstring += room.roomnumber + ";"
        arrivalamount += room.amount

    sheet.cell(row=35, column=1).value = arrivalamount
    sheet.cell(row=35, column=3).value = arrivalstring

    sheet.cell(row=35, column=5).value = departurestring
    sheet.cell(row=35, column=7).value = departureamount

    wb.save(r"transfer.xlsx")

    # store the file in in a variable
    if fileexists:
        client.collection("sheet").update(
            result[0].id,
            {
                "excelfile": FileUpload(
                    (
                        querystring,
                        open(path, "rb"),
                    )
                )
            },
        )
        print("File updated")
    else:
        client.records.create(
            "sheet",
            {
                "date": querystring,
                "excelfile": FileUpload(
                    (
                        querystring,
                        open(path, "rb"),
                    )
                ),
            },
        )

        print("File created")
